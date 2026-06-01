# -*- coding: utf-8 -*-
"""
ICMDA 숙소 배정 - 마스터 데이터 통합 스크립트

원본 3개 파일을 읽어 손으로 수정 가능한 단일 엑셀(ICMDA Master Data.xlsx)을 만든다.
탭 2개: '참가자' / '객실'.  헤더는 모두 한국어.

원본 파일은 건드리지 않는다. (읽기 전용)

데이터 우선순위:
  - 참가자 정보의 기준(spine)은 '02 참가자 데이터.xlsx'의 FINAL 시트(634행).
  - FINAL 값이 우선. 비어 있을 때만 'Final Allocation.xlsx'의 숙소배치 탭에서 보충.
"""

import datetime
import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PART_SRC = "02 참가자 데이터.xlsx"
ROOM_SRC = "01 호텔방 데이터.xlsx"
ALLOC_SRC = "Final Allocation.xlsx"
OUT = "ICMDA Master Data.xlsx"

NIGHT_COLS = list(range(15, 25))  # FINAL 열 15~24 = 6/26 ~ 7/5
NIGHT_LABELS = ["6월26일", "6월27일", "6월28일", "6월29일", "6월30일",
                "7월1일", "7월2일", "7월3일", "7월4일", "7월5일"]


def norm_name(s):
    return "".join(str(s or "").lower().split())


def to_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def to_date(v):
    """생년월일 등 날짜는 시간 없이 YYYY-MM-DD 로만."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return to_text(v)


VOWELS = set("AEIOU")

# 모음이 있어 일반 단어처럼 보이지만 약어인 것들(대문자 유지). 데이터에서 추출.
# 모음이 없는 대문자(CMF, UKZN, MD 등)는 규칙으로 자동 유지하므로 목록에 없어도 됨.
ACRONYMS = {
    "UK", "USA", "US", "UAE", "ENT",
    "ICMDA", "CMDA", "UCMF", "AMCT", "AMPCN", "AMPCI", "AMDCT", "AMDCB",
    "AMDCC", "AMDC", "ACSAI", "ACSA", "ACASAI", "GMMA", "JCMA", "EHA",
    "CMFSA", "RCAF", "IPGDFM", "CMDFI", "NUST", "BCMDA", "HCFI", "APROVIGB",
    "UPAP", "APSA", "COMAHS", "CMED", "ACUPS", "ACBAS", "KCMDA", "BELCMDA",
    "ASOME", "TSMU", "NUSOM", "PAACS", "CEML", "NCMDA", "ECUSTA", "HLI",
    "ACMEH", "NTUT", "TLMI", "CMAB", "IMC", "OMF", "AEO", "CEO", "DICO",
    "DICOS", "KNU", "AMC", "MENA", "SIM", "GCMDF", "KCMC", "KCMH",
}


def _fix_word(w):
    if w.isupper():
        core = "".join(ch for ch in w if ch.isalpha())
        if core in ACRONYMS or not (set(core) & VOWELS):
            return w                      # 약어 -> 대문자 유지
        return w[:1] + w[1:].lower()      # 일반 단어 -> 첫 글자만 대문자
    if w.islower():
        return w[:1].upper() + w[1:]
    return w                              # 혼합 대소문자(McDonald) -> 유지


def normcap(v):
    """국적/소속/직책 대소문자 정규화.
    - 소문자 단어 -> 첫 글자 대문자
    - 전부 대문자 단어 -> 첫 글자만 대문자(MALAYSIA->Malaysia),
      단 약어(ICMDA, CMF, 모음 없는 토큰)는 그대로 둔다
    - 혼합 대소문자(McDonald, Chrétien) -> 유지"""
    s = to_text(v)
    if not s:
        return ""
    return " ".join(_fix_word(w) for w in s.split())


# 한글이름 칸에 잘못 들어간 값 정리
NATIVE = re.compile(r"[가-힣぀-ヿ一-鿿]")  # 한글/가나/한자
KR_JUNK = re.compile(
    r"^(n/?a|no|none|nil|x|-|\.|not applicable|not korean|non[ -]?korean.*"
    r"|neant|n[ãa]o|kh[oô]ng c[oó])$", re.I)
KR_SEC = re.compile(r"security|reason", re.I)


def fix_korean_name(raw, en):
    """한글이름 칸 정리. 반환: (정리된 값, 검토필요?)
    - 한글/한자/가나 등 고유 이름 -> 유지
    - 쓰레기값(N/A, No 등) / 이름과 완전 중복 -> 삭제
    - 그 외 영문 조각 -> 값은 남기되 '검토필요' 표시
    """
    kr = to_text(raw)
    if not kr:
        return "", ""
    if NATIVE.search(kr):
        return kr, ""
    if KR_JUNK.match(kr) or KR_SEC.search(kr):
        return "", ""
    if kr.lower() == to_text(en).lower():
        return "", ""
    return kr, "한글이름 확인"


# ---------------------------------------------------------------- 숙소배치 보충용
def load_alloc_lookup():
    """숙소배치 탭에서 이름 -> (숙소명, 도착, 출발) 보충 정보."""
    wb = openpyxl.load_workbook(ALLOC_SRC, data_only=True)
    ws = wb["숙소배치"]
    lookup = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        cat = to_text(row[0])
        name = to_text(row[2]) if len(row) > 2 else ""
        if not name or cat in ("", "카테고리"):
            continue
        lodge = to_text(row[8]) if len(row) > 8 else ""
        arrive = to_text(row[5]) if len(row) > 5 else ""
        depart = to_text(row[6]) if len(row) > 6 else ""
        lookup[norm_name(name)] = (lodge, arrive, depart)
    wb.close()
    return lookup


# ---------------------------------------------------------------- 참가자 탭
PART_HEADERS = [
    "참가자ID", "그룹ID", "카테고리", "구분", "이름", "한글이름", "성별",
    "생년월일", "국적", "이메일", "소속", "직책", "WB등급",
    "결제상태", "배정필요",
    "도착일", "출발일",
] + NIGHT_LABELS + [
    "숙박일수", "본인", "배우자", "자녀", "기타", "총인원", "동반자요약",
    "배정숙소", "중요사항", "비고", "출처", "검토필요",
]


def build_participants():
    alloc = load_alloc_lookup()
    wb = openpyxl.load_workbook(PART_SRC, data_only=True)
    ws = wb["FINAL"]
    rows = [r for r in ws.iter_rows(min_row=5, values_only=True)
            if any(c not in (None, "") for c in r)]

    out, fam_counter = [], 0
    fam_seen = {}
    for idx, r in enumerate(rows, 1):
        def g(col):  # 1-based FINAL 열
            return r[col - 1] if col - 1 < len(r) else None

        cat = to_text(g(1))
        name = to_text(g(3)).upper()  # 영문 이름은 전부 대문자
        family_flag = to_text(g(5)).lower() == "yes"
        total = to_text(g(30)) or "1"

        # 그룹ID: 가족 단위 -> G-fam-NN, 중국팀 -> 중국팀, 그 외 빈칸
        group_id = ""
        if family_flag or (total.isdigit() and int(total) > 1):
            key = norm_name(name)
            if key not in fam_seen:
                fam_counter += 1
                fam_seen[key] = f"G-fam-{fam_counter:02d}"
            group_id = fam_seen[key]
        elif cat == "중국":
            group_id = "중국팀"

        # 숙박 야간(✓) 처리
        nights = ["✓" if g(c) not in (None, "") else "" for c in NIGHT_COLS]
        nnights = sum(1 for n in nights if n)
        nights_label = f"{nnights}박" if nnights else ""

        arrive, depart = to_text(g(10)), to_text(g(11))
        lodge = to_text(g(13))

        # FINAL 비어있을 때만 숙소배치에서 보충
        sup = alloc.get(norm_name(name))
        source = "FINAL"
        if sup:
            if not lodge and sup[0]:
                lodge = sup[0]; source = "FINAL+숙소배치"
            if not arrive and sup[1]:
                arrive = sup[1]; source = "FINAL+숙소배치"
            if not depart and sup[2]:
                depart = sup[2]; source = "FINAL+숙소배치"

        # 도착/출발이 비어 있으면 ✓ 범위로 추정
        if not arrive and any(nights):
            arrive = NIGHT_LABELS[nights.index("✓")]
        if not depart and any(nights):
            last = max(i for i, n in enumerate(nights) if n)
            depart = NIGHT_LABELS[last]

        kr_name, review = fix_korean_name(g(40), name)

        out.append([
            f"P{idx:04d}", group_id, cat, to_text(g(2)), name, kr_name,
            to_text(g(4)), to_date(g(41)), normcap(g(6)), to_text(g(45)).lower(),
            normcap(g(42)), normcap(g(43)), to_text(g(44)),
            to_text(g(7)), to_text(g(8)),
            arrive, depart, *nights, nights_label,
            to_text(g(25)), to_text(g(26)), to_text(g(27)), to_text(g(28)),
            total, to_text(g(31)), lodge, to_text(g(9)), to_text(g(12)), source,
            review,
        ])
    wb.close()
    return out


# ---------------------------------------------------------------- 객실 탭
ROOM_HEADERS = ["객실ID", "숙소", "동", "층", "호실", "객실유형",
                "성별구분", "정원", "이용가능기간", "필수온돌", "비고"]

ONDOL_CAP = {"Korean Traditional": 6, "Korean Room": 6, "BED": 8, "Bed for 2": 2}
MANDATORY_ONDOL = {117, 118, 119, 120, 121, 122, 123, 124, 125, 126, 127, 133, 134}


def build_rooms():
    wb = openpyxl.load_workbook(ROOM_SRC, data_only=True)
    out, rid = [], 0

    # --- BYYTC (부영청소년수련원) ---
    ws = wb["BYYTC ROOM 배정"]
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        rno = row[3]
        if rno in (None, ""):
            continue
        rtype = to_text(row[4])
        cap = ONDOL_CAP.get(rtype, 6)
        try:
            room_int = int(float(rno))
        except (TypeError, ValueError):
            room_int = None
        rid += 1
        out.append([
            f"R{rid:03d}", "부영청소년수련원", to_text(row[1]), to_text(row[2]),
            to_text(rno), rtype, to_text(row[5]), cap, "6/29~7/5",
            "Y" if room_int in MANDATORY_ONDOL else "", "",
        ])

    # --- YOUUS HOTEL (유어스호텔) ---
    ws2 = wb["YOUUS HOTEL 배정"]
    for row in ws2.iter_rows(min_row=5, max_row=ws2.max_row, values_only=True):
        rtype = to_text(row[3])
        if not rtype or "Type" in rtype:
            continue
        total = row[6]
        try:
            cap = int(float(total))
        except (TypeError, ValueError):
            cap = 2
        rid += 1
        out.append([
            f"R{rid:03d}", "유어스호텔", to_text(row[0]), to_text(row[1]),
            "", rtype, "", cap, "6/29~7/6", "", "",
        ])
    wb.close()
    return out


# ---------------------------------------------------------------- 쓰기 + 서식
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 카테고리(그룹 유형)별 행 배경색 — 구분이 쉽도록
CATEGORY_FILLS = {
    "무료숙소신청 - 개인": "E8F1FB",  # 연파랑
    "무료숙소신청 - 가족": "E2EFDA",  # 연초록
    "NOC 위원": "FFF2CC",            # 연노랑
    "AFF": "FCE4D6",                 # 연주황
    "중국": "FADBD8",                # 연빨강
}


def write_sheet(ws, headers, rows, cat_col=None):
    ws.append(headers)
    for r in rows:
        ws.append(r)

    # 카테고리별 행 색칠
    if cat_col is not None:
        for ri, r in enumerate(rows, start=2):
            color = CATEGORY_FILLS.get(to_text(r[cat_col - 1]))
            if not color:
                continue
            fill = PatternFill("solid", fgColor=color)
            for j in range(1, len(headers) + 1):
                ws.cell(row=ri, column=j).fill = fill
    for j, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=j)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        # 폭: 헤더/내용 최대 길이 기준 (한글 가중치)
        maxlen = len(str(headers[j - 1]))
        for r in rows[:200]:
            v = str(r[j - 1]) if j - 1 < len(r) else ""
            maxlen = max(maxlen, len(v))
        ws.column_dimensions[get_column_letter(j)].width = min(max(maxlen + 2, 6), 30)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def main():
    parts = build_participants()
    rooms = build_rooms()

    wb = openpyxl.Workbook()
    write_sheet(wb.active, PART_HEADERS, parts,
                cat_col=PART_HEADERS.index("카테고리") + 1)
    wb.active.title = "참가자"
    write_sheet(wb.create_sheet("객실"), ROOM_HEADERS, rooms)
    wb.save(OUT)

    print(f"생성 완료: {OUT}")
    print(f"  참가자: {len(parts)}행")
    print(f"  객실:   {len(rooms)}행, 총 정원 {sum(r[7] for r in rooms)}")


if __name__ == "__main__":
    main()
